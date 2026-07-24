from __future__ import annotations

import importlib
import re
from collections.abc import Callable
from pathlib import Path
from typing import Any, cast
from xml.etree import ElementTree
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from app.schemas.report import TemplateParseResult
from app.services.template.base import TemplateService
from app.services.template.render import extract_placeholders, extract_raw_placeholders

PLACEHOLDER_PATTERN = re.compile(r"\{\{\s*([^{}\r\n]+?)\s*\}\}")
PDF_TEXT_LITERAL_PATTERN = re.compile(r"\((?P<value>(?:\\.|[^\\()])*)\)\s*Tj")
PDF_TEXT_ARRAY_PATTERN = re.compile(
    r"\[(?P<value>(?:\s*\((?:\\.|[^\\()])*\)\s*-?\d*\.?\d*)+)\s*\]\s*TJ",
    re.DOTALL,
)
PDF_ARRAY_LITERAL_PATTERN = re.compile(r"\((?P<value>(?:\\.|[^\\()])*)\)")
PDF_TEXT_BLOCK_PATTERN = re.compile(r"BT(?P<value>.*?)ET", re.DOTALL)
MAX_PDF_PARSE_PAGES = 10
MAX_PDF_FALLBACK_BYTES = 4 * 1024 * 1024
PdfOpen = Callable[[str], Any]


class DocxTemplateService(TemplateService):
    async def parse_template(self, file_path: str) -> TemplateParseResult:
        template_path = Path(file_path)
        suffix = template_path.suffix.lower()
        if suffix not in {".docx", ".xlsx", ".pdf"}:
            msg = "Only .docx, .xlsx and .pdf template files are supported now."
            raise ValueError(msg)
        if not template_path.exists():
            msg = f"Template file not found: {file_path}"
            raise FileNotFoundError(msg)

        template_text = read_template_text(template_path)
        fields = extract_placeholders(template_text)
        template_type = _guess_template_type(fields, template_path.name)

        return TemplateParseResult(
            template_type=template_type,
            fields=fields,
            description=f"Parsed {len(fields)} placeholder(s) from {template_path.name}",
            raw_content={
                "source": suffix.removeprefix("."),
                "path": str(template_path),
                "template_text": template_text,
                "raw_placeholders": extract_raw_placeholders(template_text),
                "placeholder_count": len(fields),
            },
        )


def read_template_text(template_path: Path) -> str:
    suffix = template_path.suffix.lower()
    if suffix == ".docx":
        return _read_docx_text(template_path)
    if suffix == ".xlsx":
        return _read_xlsx_text(template_path)
    if suffix == ".pdf":
        return _read_pdf_text(template_path)
    return ""


def _read_docx_text(template_path: Path) -> str:
    try:
        with ZipFile(template_path) as docx:
            xml_names = [
                name
                for name in docx.namelist()
                if name == "word/document.xml"
                or name.startswith(("word/header", "word/footer"))
            ]
            return "\n".join(
                _extract_xml_paragraphs(docx.read(name).decode("utf-8")) for name in xml_names
            )
    except BadZipFile as exc:
        msg = f"Invalid docx template file: {template_path}"
        raise ValueError(msg) from exc


def _extract_xml_paragraphs(xml_text: str) -> str:
    root = ElementTree.fromstring(xml_text)
    paragraphs: list[str] = []
    for paragraph in root.iter():
        if not (paragraph.tag.endswith("}p") or paragraph.tag == "p"):
            continue
        text = "".join(
            element.text or ""
            for element in paragraph.iter()
            if element.tag.endswith("}t") or element.tag == "t"
        ).strip()
        if text:
            paragraphs.append(text)
    return "\n".join(paragraphs)


def _read_xlsx_text(template_path: Path) -> str:
    workbook = load_workbook(template_path, data_only=False)
    text_parts: list[str] = []
    for worksheet in workbook.worksheets:
        for row in worksheet.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    text_parts.append(cell.value)
    return "\n".join(text_parts)


def _read_pdf_text(template_path: Path) -> str:
    extracted = _read_pdf_text_with_pymupdf(template_path)
    if extracted.strip():
        return extracted
    return _read_pdf_text_fallback(template_path)


def _read_pdf_text_with_pymupdf(template_path: Path) -> str:
    open_pdf = _import_pymupdf_open()
    if open_pdf is None:
        return ""

    document = None
    try:
        document = open_pdf(str(template_path))
        text_parts: list[str] = []
        for page_index in range(min(len(document), MAX_PDF_PARSE_PAGES)):
            page = document[page_index]
            get_text = getattr(page, "get_text", None)
            if not callable(get_text):
                continue
            page_text = get_text("text")
            if isinstance(page_text, str) and page_text.strip():
                text_parts.append(page_text.strip())
        return "\n".join(text_parts)
    except Exception:
        return ""
    finally:
        close = getattr(document, "close", None)
        if callable(close):
            close()


def _import_pymupdf_open() -> PdfOpen | None:
    for module_name in ("pymupdf", "fitz"):
        try:
            module = importlib.import_module(module_name)
        except ImportError:
            continue
        open_pdf = getattr(module, "open", None)
        if callable(open_pdf):
            return cast(PdfOpen, open_pdf)
    return None


def _read_pdf_text_fallback(template_path: Path) -> str:
    raw_bytes = template_path.read_bytes()[:MAX_PDF_FALLBACK_BYTES]
    raw_text = raw_bytes.decode("latin-1", errors="ignore")
    if "{{" not in raw_text and "Tj" not in raw_text and "TJ" not in raw_text:
        return ""

    text_parts: list[str] = []
    for block_match in PDF_TEXT_BLOCK_PATTERN.finditer(raw_text):
        block_text = block_match.group("value")
        text_parts.extend(_pdf_literals_from_text_block(block_text))
    if text_parts:
        return "\n".join(part for part in text_parts if part.strip())

    placeholders = PLACEHOLDER_PATTERN.findall(raw_text)
    return "\n".join(f"{{{{{placeholder.strip()}}}}}" for placeholder in placeholders)


def _pdf_literals_from_text_block(block_text: str) -> list[str]:
    parts = [
        _decode_pdf_literal(match.group("value"))
        for match in PDF_TEXT_LITERAL_PATTERN.finditer(block_text)
    ]
    for array_match in PDF_TEXT_ARRAY_PATTERN.finditer(block_text):
        text = "".join(
            _decode_pdf_literal(match.group("value"))
            for match in PDF_ARRAY_LITERAL_PATTERN.finditer(array_match.group("value"))
        )
        if text:
            parts.append(text)
    return parts


def _decode_pdf_literal(value: str) -> str:
    replacements = {
        "n": "\n",
        "r": "\r",
        "t": "\t",
        "b": "\b",
        "f": "\f",
        "(": "(",
        ")": ")",
        "\\": "\\",
    }
    decoded: list[str] = []
    index = 0
    while index < len(value):
        char = value[index]
        if char != "\\":
            decoded.append(char)
            index += 1
            continue
        index += 1
        if index >= len(value):
            break
        escaped = value[index]
        if escaped in "\r\n":
            index += 1
            if escaped == "\r" and index < len(value) and value[index] == "\n":
                index += 1
            continue
        if escaped in "01234567":
            octal = escaped
            index += 1
            for _ in range(2):
                if index >= len(value) or value[index] not in "01234567":
                    break
                octal += value[index]
                index += 1
            decoded.append(chr(int(octal, 8)))
            continue
        decoded.append(replacements.get(escaped, escaped))
        index += 1
    return "".join(decoded)


def _guess_template_type(fields: list[str], file_name: str) -> str:
    lowered_name = file_name.lower()
    if "weekly" in lowered_name or "week" in lowered_name or "周报" in file_name:
        return "weekly"
    if "daily" in lowered_name or "day" in lowered_name or "日报" in file_name:
        return "daily"
    if "week_summary" in fields or "next_week_plan" in fields:
        return "weekly"
    return "daily"
