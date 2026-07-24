from __future__ import annotations

import base64
import html
import logging
import re
from pathlib import Path
from typing import Any

from docx import Document
from docx.oxml.ns import qn
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook
from openpyxl.cell.cell import MergedCell
from openpyxl.utils import get_column_letter
from openpyxl.utils.cell import range_boundaries

logger = logging.getLogger(__name__)

PLACEHOLDER_PATTERN = re.compile(r"(\{\{\s*[^{}\r\n]+?\s*\}\})")
MAX_XLSX_ROWS = 80
MAX_XLSX_COLUMNS = 30
MAX_PDF_PREVIEW_BYTES = 8 * 1024 * 1024


def build_source_preview_html(file_path: str | None) -> str | None:
    if not file_path:
        return None

    template_path = Path(file_path)
    if not template_path.exists():
        return None

    try:
        if template_path.suffix.lower() == ".docx":
            return _docx_preview_html(template_path)
        if template_path.suffix.lower() == ".xlsx":
            return _xlsx_preview_html(template_path)
        if template_path.suffix.lower() == ".pdf":
            return _pdf_preview_html(template_path)
    except Exception:
        logger.exception("Failed to build template source preview for %s", template_path)
        return None
    return None


def _docx_preview_html(template_path: Path) -> str:
    document = Document(str(template_path))
    blocks = [_docx_block_html(block) for block in _iter_docx_blocks(document)]
    content = "".join(block for block in blocks if block)
    if not content:
        content = '<p class="source-empty">空白模板</p>'
    return f'<div class="source-preview source-preview-docx">{content}</div>'


def _iter_docx_blocks(document: Any) -> list[Any]:
    blocks: list[Any] = []
    body = document.element.body
    for child in body.iterchildren():
        if child.tag.endswith("}p"):
            blocks.append(Paragraph(child, document))
        elif child.tag.endswith("}tbl"):
            blocks.append(Table(child, document))
    return blocks


def _docx_block_html(block: Any) -> str:
    if isinstance(block, Paragraph):
        return _docx_paragraph_html(block)
    if isinstance(block, Table):
        return _docx_table_html(block)
    return ""


def _docx_paragraph_html(paragraph: Any) -> str:
    runs = [_docx_run_html(run) for run in paragraph.runs]
    content = "".join(runs) or "&nbsp;"
    style_attr = _style_attribute(_docx_paragraph_styles(paragraph))
    class_attr = _docx_paragraph_class(paragraph)
    return f"<p{class_attr}{style_attr}>{content}</p>"


def _docx_paragraph_class(paragraph: Any) -> str:
    style_name = _safe_docx_style_name(paragraph)
    if not style_name:
        return ""
    lowered = style_name.lower()
    if "title" in lowered:
        return ' class="source-heading source-title"'
    if "heading" in lowered or "标题" in style_name:
        return ' class="source-heading"'
    return ""


def _safe_docx_style_name(paragraph: Any) -> str | None:
    style = getattr(paragraph, "style", None)
    name = getattr(style, "name", None)
    return name if isinstance(name, str) else None


def _docx_paragraph_styles(paragraph: Any) -> list[tuple[str, str | None]]:
    alignment_map = {
        "CENTER": "center",
        "RIGHT": "right",
        "JUSTIFY": "justify",
        "LEFT": "left",
    }
    alignment = getattr(paragraph, "alignment", None)
    alignment_name = getattr(alignment, "name", None)
    text_align = alignment_map.get(alignment_name) if isinstance(alignment_name, str) else None
    return [("text-align", text_align)]


def _docx_run_html(run: Any) -> str:
    text = getattr(run, "text", "")
    if not isinstance(text, str) or not text:
        return ""
    style_attr = _style_attribute(_docx_run_styles(run))
    return f"<span{style_attr}>{_text_with_placeholders(text)}</span>"


def _docx_run_styles(run: Any) -> list[tuple[str, str | None]]:
    font = run.font
    size = getattr(font, "size", None)
    color = getattr(getattr(font, "color", None), "rgb", None)
    highlight = getattr(font, "highlight_color", None)
    return [
        ("font-family", _docx_font_family(run)),
        ("font-size", f"{size.pt:.1f}pt" if size is not None else None),
        ("font-weight", "700" if bool(getattr(run, "bold", False)) else None),
        ("font-style", "italic" if bool(getattr(run, "italic", False)) else None),
        ("text-decoration", "underline" if bool(getattr(run, "underline", False)) else None),
        ("color", f"#{color}" if color is not None else None),
        ("background-color", _docx_highlight_color(highlight)),
    ]


def _docx_font_family(run: Any) -> str | None:
    font_name = getattr(run.font, "name", None)
    element = getattr(run, "_element", None)
    rpr = getattr(element, "rPr", None)
    rfonts = getattr(rpr, "rFonts", None)
    if rfonts is not None:
        for key in ("w:eastAsia", "w:ascii", "w:hAnsi"):
            value = rfonts.get(qn(key))
            if isinstance(value, str) and value.strip():
                return value
    return font_name if isinstance(font_name, str) and font_name.strip() else None


def _docx_highlight_color(highlight: Any) -> str | None:
    if highlight is None:
        return None
    colors = {
        "YELLOW": "#fff59d",
        "GREEN": "#bbf7d0",
        "CYAN": "#bae6fd",
        "MAGENTA": "#fbcfe8",
        "BLUE": "#bfdbfe",
        "RED": "#fecaca",
        "DARK_YELLOW": "#fde68a",
        "DARK_GREEN": "#86efac",
        "DARK_CYAN": "#67e8f9",
        "DARK_MAGENTA": "#f9a8d4",
        "DARK_BLUE": "#93c5fd",
        "DARK_RED": "#fca5a5",
        "BLACK": "#111827",
        "GRAY_25": "#f3f4f6",
        "GRAY_50": "#e5e7eb",
    }
    name = getattr(highlight, "name", None)
    return colors.get(name) if isinstance(name, str) else None


def _docx_table_html(table: Any) -> str:
    rows: list[str] = []
    for row in table.rows:
        cells = []
        for cell in row.cells:
            content = "".join(_docx_paragraph_html(paragraph) for paragraph in cell.paragraphs)
            cells.append(f"<td>{content or '&nbsp;'}</td>")
        rows.append(f"<tr>{''.join(cells)}</tr>")
    return f'<table class="source-table"><tbody>{"".join(rows)}</tbody></table>'


def _xlsx_preview_html(template_path: Path) -> str:
    workbook = load_workbook(template_path, data_only=False)
    sheet_html = [_worksheet_preview_html(worksheet) for worksheet in workbook.worksheets]
    content = "".join(sheet_html) or '<p class="source-empty">空白模板</p>'
    return f'<div class="source-preview source-preview-xlsx">{content}</div>'


def _pdf_preview_html(template_path: Path) -> str | None:
    if template_path.stat().st_size > MAX_PDF_PREVIEW_BYTES:
        return None
    encoded_pdf = base64.b64encode(template_path.read_bytes()).decode("ascii")
    return (
        '<div class="source-preview source-preview-pdf">'
        f'<object data="data:application/pdf;base64,{encoded_pdf}" type="application/pdf">'
        '<p class="source-empty">PDF 预览不可用</p>'
        "</object>"
        "</div>"
    )


def _worksheet_preview_html(worksheet: Any) -> str:
    title = html.escape(str(worksheet.title))
    max_row = min(int(worksheet.max_row or 1), MAX_XLSX_ROWS)
    max_column = min(int(worksheet.max_column or 1), MAX_XLSX_COLUMNS)
    merge_spans, covered_cells = _xlsx_merge_maps(worksheet, max_row, max_column)

    colgroup = "".join(_xlsx_col_html(worksheet, column) for column in range(1, max_column + 1))
    rows: list[str] = []
    for row_index in range(1, max_row + 1):
        cells: list[str] = []
        row_style = _style_attribute(_xlsx_row_styles(worksheet, row_index))
        for column_index in range(1, max_column + 1):
            if (row_index, column_index) in covered_cells:
                continue
            cell = worksheet.cell(row=row_index, column=column_index)
            if isinstance(cell, MergedCell):
                continue
            rowspan, colspan = merge_spans.get((row_index, column_index), (1, 1))
            cells.append(_xlsx_cell_html(cell, rowspan, colspan))
        rows.append(f"<tr{row_style}>{''.join(cells)}</tr>")

    clipped = worksheet.max_row > max_row or worksheet.max_column > max_column
    note = (
        f'<p class="source-note">仅显示前 {max_row} 行、{max_column} 列</p>'
        if clipped
        else ""
    )
    return (
        '<section class="source-sheet">'
        f"<h4>{title}</h4>"
        f"{note}"
        f'<table class="source-table"><colgroup>{colgroup}</colgroup>'
        f'<tbody>{"".join(rows)}</tbody></table>'
        "</section>"
    )


def _xlsx_col_html(worksheet: Any, column: int) -> str:
    dimension = worksheet.column_dimensions[get_column_letter(column)]
    width = getattr(dimension, "width", None)
    if isinstance(width, int | float):
        pixels = max(48, min(240, int(width * 7)))
        return f'<col style="width:{pixels}px">'
    return '<col style="width:96px">'


def _xlsx_merge_maps(
    worksheet: Any,
    max_row: int,
    max_column: int,
) -> tuple[dict[tuple[int, int], tuple[int, int]], set[tuple[int, int]]]:
    spans: dict[tuple[int, int], tuple[int, int]] = {}
    covered: set[tuple[int, int]] = set()
    for merged_range in worksheet.merged_cells.ranges:
        min_column, min_row, range_max_column, range_max_row = range_boundaries(str(merged_range))
        if min_row > max_row or min_column > max_column:
            continue
        clipped_max_row = min(range_max_row, max_row)
        clipped_max_column = min(range_max_column, max_column)
        spans[(min_row, min_column)] = (
            clipped_max_row - min_row + 1,
            clipped_max_column - min_column + 1,
        )
        for row in range(min_row, clipped_max_row + 1):
            for column in range(min_column, clipped_max_column + 1):
                if (row, column) != (min_row, min_column):
                    covered.add((row, column))
    return spans, covered


def _xlsx_row_styles(worksheet: Any, row_index: int) -> list[tuple[str, str | None]]:
    height = worksheet.row_dimensions[row_index].height
    return [("height", f"{height:.1f}pt" if isinstance(height, int | float) else None)]


def _xlsx_cell_html(cell: Any, rowspan: int, colspan: int) -> str:
    value = "" if cell.value is None else str(cell.value)
    attrs = [
        f'rowspan="{rowspan}"' if rowspan > 1 else "",
        f'colspan="{colspan}"' if colspan > 1 else "",
        _style_attribute(_xlsx_cell_styles(cell)).strip(),
    ]
    attr_text = " ".join(attr for attr in attrs if attr)
    if attr_text:
        attr_text = f" {attr_text}"
    content = _text_with_placeholders(value) if value else "&nbsp;"
    return f"<td{attr_text}>{content}</td>"


def _xlsx_cell_styles(cell: Any) -> list[tuple[str, str | None]]:
    font = cell.font
    fill = cell.fill
    alignment = cell.alignment
    border = cell.border
    fill_color = _openpyxl_color(getattr(fill, "fgColor", None)) if fill.fill_type else None
    border_color = _xlsx_border_color(border)
    return [
        ("font-family", font.name if isinstance(font.name, str) else None),
        ("font-size", f"{font.sz:.1f}pt" if isinstance(font.sz, int | float) else None),
        ("font-weight", "700" if font.bold else None),
        ("font-style", "italic" if font.italic else None),
        ("text-decoration", "underline" if font.underline else None),
        ("color", _openpyxl_color(font.color)),
        ("background-color", fill_color),
        ("text-align", alignment.horizontal if isinstance(alignment.horizontal, str) else None),
        ("vertical-align", alignment.vertical if isinstance(alignment.vertical, str) else None),
        ("white-space", "pre-wrap" if alignment.wrap_text else None),
        ("border-color", border_color),
    ]


def _xlsx_border_color(border: Any) -> str | None:
    for side in (border.left, border.right, border.top, border.bottom):
        if side.style:
            color = _openpyxl_color(side.color)
            return color or "#d0d5dd"
    return None


def _openpyxl_color(color: Any) -> str | None:
    if color is None:
        return None
    color_type = getattr(color, "type", None)
    if color_type != "rgb":
        return None
    rgb = getattr(color, "rgb", None)
    if not isinstance(rgb, str) or len(rgb) < 6:
        return None
    value = rgb[-6:].upper()
    if not re.fullmatch(r"[0-9A-F]{6}", value):
        return None
    return f"#{value}"


def _text_with_placeholders(text: str) -> str:
    parts: list[str] = []
    placeholder_re = re.compile(r"^\{\{\s*[^{}\r\n]+?\s*\}\}$")
    for part in PLACEHOLDER_PATTERN.split(text):
        if not part:
            continue
        escaped = html.escape(part).replace("\n", "<br>").replace("\t", "&emsp;")
        if placeholder_re.fullmatch(part):
            parts.append(f'<mark class="source-placeholder">{escaped}</mark>')
        else:
            parts.append(escaped)
    return "".join(parts)


def _style_attribute(styles: list[tuple[str, str | None]]) -> str:
    cleaned = []
    for name, value in styles:
        clean_value = _clean_css_value(value)
        if clean_value:
            cleaned.append(f"{name}:{clean_value}")
    if not cleaned:
        return ""
    return f' style="{html.escape("; ".join(cleaned), quote=True)}"'


def _clean_css_value(value: str | None) -> str | None:
    if not value:
        return None
    cleaned = str(value).strip()
    for char in (";", "<", ">", '"', "'", "{", "}", "\\"):
        cleaned = cleaned.replace(char, "")
    return cleaned or None
