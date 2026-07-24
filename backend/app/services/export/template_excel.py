from __future__ import annotations

import re
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook

from app.core.config import get_settings
from app.schemas.report import ExportResult, ReportContent
from app.services.export.base import ExportService
from app.services.template.render import PLACEHOLDER_PATTERN, build_replacements


class TemplateExcelExportService(ExportService):
    async def export(self, report: ReportContent, export_type: str) -> ExportResult:
        return await self.export_with_template(report, export_type, None)

    async def export_with_template(
        self,
        report: ReportContent,
        export_type: str,
        template_path: str | None,
    ) -> ExportResult:
        if export_type.lower() not in {"xlsx", "excel"}:
            msg = f"Unsupported Excel export type: {export_type}"
            raise ValueError(msg)
        if not template_path:
            msg = "template_path is required for template Excel export."
            raise ValueError(msg)

        source_template = Path(template_path)
        if source_template.suffix.lower() != ".xlsx":
            msg = "Only .xlsx template files are supported now."
            raise ValueError(msg)
        if not source_template.exists():
            msg = f"Template file not found: {template_path}"
            raise FileNotFoundError(msg)

        settings = get_settings()
        export_dir = Path(settings.storage_root) / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        safe_title = re.sub(r"[^A-Za-z0-9._\-\u4e00-\u9fff]+", "_", report.title).strip("_")
        target_path = export_dir / f"{safe_title or 'report'}-template-{uuid4().hex[:8]}.xlsx"

        replacements = build_replacements(report)
        workbook = load_workbook(source_template)
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str):
                        cell.value = PLACEHOLDER_PATTERN.sub(
                            lambda match: replacements.get(match.group(1), match.group(0)),
                            cell.value,
                        )
        workbook.save(target_path)

        return ExportResult(
            export_type="xlsx",
            file_path=str(target_path),
            status="success",
            download_url=f"/storage/exports/{target_path.name}",
        )
