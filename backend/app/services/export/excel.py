from __future__ import annotations

import re
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet

from app.core.config import get_settings
from app.schemas.report import ExportResult, ReportContent, TaskItem
from app.services.export.base import ExportService


class ExcelExportService(ExportService):
    async def export(self, report: ReportContent, export_type: str) -> ExportResult:
        if export_type.lower() not in {"xlsx", "excel"}:
            msg = f"Unsupported Excel export type: {export_type}"
            raise ValueError(msg)

        settings = get_settings()
        export_dir = Path(settings.storage_root) / "exports"
        export_dir.mkdir(parents=True, exist_ok=True)

        safe_title = re.sub(r"[^A-Za-z0-9._\-\u4e00-\u9fff]+", "_", report.title).strip("_")
        file_path = export_dir / f"{safe_title or 'report'}-{uuid4().hex[:8]}.xlsx"

        workbook = _build_workbook(report)
        workbook.save(file_path)

        return ExportResult(
            export_type="xlsx",
            file_path=str(file_path),
            status="success",
            download_url=f"/storage/exports/{file_path.name}",
        )


def _build_workbook(report: ReportContent) -> Workbook:
    workbook = Workbook()
    overview = workbook.active
    overview.title = "报表概览"
    _fill_overview_sheet(overview, report)

    tasks = workbook.create_sheet("任务明细")
    _fill_tasks_sheet(tasks, report)

    return workbook


def _fill_overview_sheet(sheet: Worksheet, report: ReportContent) -> None:
    rows = [
        ("标题", report.title),
        ("报表类型", _report_type_label(report.report_type)),
        ("日期", report.date.isoformat()),
        ("风格", report.style),
        ("工作总结", report.summary),
        ("问题与风险", _format_list(report.problems)),
        ("解决方案", _format_list(report.solutions)),
        ("下一步计划", _format_list(report.next_plan)),
    ]
    sheet.append(["字段", "内容"])
    for row in rows:
        sheet.append(list(row))

    _style_header(sheet)
    sheet.column_dimensions["A"].width = 18
    sheet.column_dimensions["B"].width = 72
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _fill_tasks_sheet(sheet: Worksheet, report: ReportContent) -> None:
    sheet.append(["分组", "任务ID", "标题", "描述", "状态", "进度", "置信度", "用户确认"])
    for group_name, tasks in [
        ("已完成", report.completed_tasks),
        ("进行中", report.in_progress_tasks),
    ]:
        if not tasks:
            sheet.append([group_name, "", "暂无", "", "", "", "", ""])
            continue
        for task in tasks:
            sheet.append(_task_row(group_name, task))

    _style_header(sheet)
    widths = [14, 18, 28, 48, 16, 12, 12, 12]
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[chr(64 + index)].width = width
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def _task_row(group_name: str, task: TaskItem) -> list[object]:
    return [
        group_name,
        task.id,
        task.title,
        task.description or "",
        task.status,
        task.progress,
        task.confidence,
        "是" if task.user_confirmed else "否",
    ]


def _style_header(sheet: Worksheet) -> None:
    fill = PatternFill("solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(vertical="center")


def _format_list(items: list[str]) -> str:
    if not items:
        return "暂无"
    return "\n".join(f"{index}. {item}" for index, item in enumerate(items, start=1))


def _report_type_label(report_type: str) -> str:
    labels = {"daily": "日报", "weekly": "周报"}
    return labels.get(report_type, report_type)
