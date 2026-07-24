from __future__ import annotations

from collections.abc import Generator
from datetime import date
from pathlib import Path
from zipfile import ZipFile

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app import models as model_registry
from app.api.v1.endpoints.reports import _build_initial_report_content
from app.core.config import get_settings
from app.core.database import Base, get_db
from app.main import app
from app.models import Report, Template, UploadedFile
from app.schemas.report import ReportCreate


def test_export_endpoint_uses_report_id_content(monkeypatch, tmp_path) -> None:
    assert model_registry.Report.__tablename__ == "reports"
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    payload = ReportCreate(
        report_type="daily",
        title="真实日报",
        report_date=date(2026, 7, 22),
        source_data={
            "content": {
                "summary": "这是数据库里的报表内容。",
                "completed_tasks": [
                    {
                        "id": "task-1",
                        "title": "完成真实 report_id 导出",
                        "description": "从 reports 表读取 content。",
                        "status": "completed",
                        "progress": 100,
                    }
                ],
                "next_plan": ["接入模板记录"],
            }
        },
    )
    content = _build_initial_report_content(payload)

    with session_factory() as session:
        stored_report = Report(
            report_type=payload.report_type,
            title=payload.title,
            report_date=payload.report_date,
            status="draft",
            content=content.model_dump(mode="json"),
            source_data=payload.source_data,
        )
        session.add(stored_report)
        session.commit()
        session.refresh(stored_report)
        report_id = stored_report.id

    template_path = tmp_path / "template.docx"
    _create_template_docx(template_path)

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        response = client.post(
            f"/api/v1/reports/{report_id}/export",
            json={"export_type": "docx", "template_path": str(template_path)},
        )
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert response.status_code == 200
    result = response.json()["data"]
    with ZipFile(result["file_path"]) as docx:
        document_xml = docx.read("word/document.xml").decode("utf-8")

    assert "真实日报" in document_xml
    assert "这是数据库里的报表内容" in document_xml
    assert "完成真实 report_id 导出" in document_xml
    assert "{{title}}" not in document_xml


def test_export_endpoint_uses_bound_docx_template(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    template_path = tmp_path / "uploads" / "custom-template.docx"
    template_path.parent.mkdir(parents=True, exist_ok=True)
    _create_template_docx(template_path)

    payload = ReportCreate(
        report_type="daily",
        title="Bound Template Report",
        report_date=date(2026, 7, 24),
        source_data={
            "content": {
                "summary": "Rendered through the bound template.",
                "completed_tasks": [
                    {
                        "id": "task-1",
                        "title": "Bind template_id to export",
                        "status": "completed",
                        "progress": 100,
                    }
                ],
                "custom_fields": {"owner": "Template Owner"},
            },
        },
    )
    content = _build_initial_report_content(payload)

    with session_factory() as session:
        file_record = UploadedFile(
            original_name="custom-template.docx",
            stored_name="custom-template.docx",
            file_type="docx",
            file_size=template_path.stat().st_size,
            storage_path=str(template_path),
        )
        session.add(file_record)
        session.flush()
        template = Template(
            name="Custom Template",
            template_type="daily",
            file_id=file_record.id,
            field_config={"fields": ["title", "summary", "completed_tasks", "owner"]},
        )
        session.add(template)
        session.flush()
        stored_report = Report(
            template_id=template.id,
            report_type=payload.report_type,
            title=payload.title,
            report_date=payload.report_date,
            status="draft",
            content=content.model_dump(mode="json"),
            source_data=payload.source_data,
        )
        session.add(stored_report)
        session.commit()
        session.refresh(stored_report)
        report_id = stored_report.id

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        get_response = client.get(f"/api/v1/reports/{report_id}")
        response = client.post(
            f"/api/v1/reports/{report_id}/export",
            json={"export_type": "docx"},
        )
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert get_response.status_code == 200
    rendered_template = get_response.json()["data"]["content"]["custom_fields"]["rendered_template"]
    assert "Bound Template Report" in rendered_template
    assert "Template Owner" in rendered_template

    assert response.status_code == 200
    result = response.json()["data"]
    assert "-template-" in Path(result["file_path"]).name
    with ZipFile(result["file_path"]) as docx:
        document_xml = docx.read("word/document.xml").decode("utf-8")

    assert "Bound Template Report" in document_xml
    assert "Rendered through the bound template." in document_xml
    assert "Bind template_id to export" in document_xml
    assert "Template Owner" in document_xml
    assert "{{title}}" not in document_xml


def test_export_endpoint_uses_bound_xlsx_template(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    template_path = tmp_path / "uploads" / "custom-template.xlsx"
    template_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet["A1"] = "{{title}}"
    worksheet["A2"] = "{{summary}}"
    worksheet["A3"] = "{{owner}}"
    workbook.save(template_path)

    payload = ReportCreate(
        report_type="daily",
        title="Bound Excel Report",
        report_date=date(2026, 7, 24),
        source_data={
            "content": {
                "summary": "Rendered into Excel template.",
                "custom_fields": {"owner": "Excel Owner"},
            },
        },
    )
    content = _build_initial_report_content(payload)

    with session_factory() as session:
        file_record = UploadedFile(
            original_name="custom-template.xlsx",
            stored_name="custom-template.xlsx",
            file_type="xlsx",
            file_size=template_path.stat().st_size,
            storage_path=str(template_path),
        )
        session.add(file_record)
        session.flush()
        template = Template(
            name="Custom Excel Template",
            template_type="daily",
            file_id=file_record.id,
            field_config={"fields": ["title", "summary", "owner"]},
        )
        session.add(template)
        session.flush()
        stored_report = Report(
            template_id=template.id,
            report_type=payload.report_type,
            title=payload.title,
            report_date=payload.report_date,
            status="draft",
            content=content.model_dump(mode="json"),
            source_data=payload.source_data,
        )
        session.add(stored_report)
        session.commit()
        session.refresh(stored_report)
        report_id = stored_report.id

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        response = client.post(
            f"/api/v1/reports/{report_id}/export",
            json={"export_type": "xlsx"},
        )
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert response.status_code == 200
    result = response.json()["data"]
    assert "-template-" in Path(result["file_path"]).name
    rendered_workbook = load_workbook(result["file_path"])
    rendered_sheet = rendered_workbook.active
    assert rendered_sheet["A1"].value == "Bound Excel Report"
    assert rendered_sheet["A2"].value == "Rendered into Excel template."
    assert rendered_sheet["A3"].value == "Excel Owner"


def test_export_endpoint_routes_xlsx_and_pdf(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    payload = ReportCreate(
        report_type="daily",
        title="多格式日报",
        report_date=date(2026, 7, 22),
        source_data={
            "content": {
                "summary": "接口应导出真实 xlsx 和 pdf 文件。",
                "completed_tasks": [
                    {
                        "id": "task-1",
                        "title": "接入多格式导出",
                        "status": "completed",
                        "progress": 100,
                    }
                ],
            }
        },
    )
    content = _build_initial_report_content(payload)

    with session_factory() as session:
        stored_report = Report(
            report_type=payload.report_type,
            title=payload.title,
            report_date=payload.report_date,
            status="draft",
            content=content.model_dump(mode="json"),
            source_data=payload.source_data,
        )
        session.add(stored_report)
        session.commit()
        session.refresh(stored_report)
        report_id = stored_report.id

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        xlsx_response = client.post(
            f"/api/v1/reports/{report_id}/export",
            json={"export_type": "xlsx"},
        )
        pdf_response = client.post(
            f"/api/v1/reports/{report_id}/export",
            json={"export_type": "pdf"},
        )
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert xlsx_response.status_code == 200
    xlsx_result = xlsx_response.json()["data"]
    assert xlsx_result["export_type"] == "xlsx"
    workbook = load_workbook(xlsx_result["file_path"])
    assert workbook["报表概览"]["B2"].value == "多格式日报"

    assert pdf_response.status_code == 200
    pdf_result = pdf_response.json()["data"]
    assert pdf_result["export_type"] == "pdf"
    pdf_content = Path(pdf_result["file_path"]).read_bytes()
    assert pdf_content.startswith(b"%PDF-")
    assert b"ReportLab" in pdf_content


def _create_template_docx(template_path) -> None:
    from app.services.export.word import _write_docx

    _write_docx(
        template_path,
        """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:p><w:r><w:t>{{title}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>{{date}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>{{summary}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>{{completed_tasks}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>{{next_plan}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>{{owner}}</w:t></w:r></w:p>
  </w:body>
</w:document>
""",
    )
