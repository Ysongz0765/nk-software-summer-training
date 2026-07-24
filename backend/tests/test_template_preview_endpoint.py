from __future__ import annotations

from collections.abc import Generator

from fastapi.testclient import TestClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from reportlab.pdfgen import canvas
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker
from sqlalchemy.pool import StaticPool

from app.core.config import get_settings
from app.core.database import Base, get_db
from app.main import app
from app.models import Template, UploadedFile
from app.services.export.word import _write_docx


def test_template_preview_endpoint_reads_uploaded_docx(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    template_path = tmp_path / "uploads" / "preview-template.docx"
    template_path.parent.mkdir(parents=True, exist_ok=True)
    _write_docx(
        template_path,
        """<?xml version="1.0" encoding="UTF-8" standalone="yes"?>
<w:document xmlns:w="http://schemas.openxmlformats.org/wordprocessingml/2006/main">
  <w:body>
    <w:p><w:r><w:t>{{title}}</w:t></w:r></w:p>
    <w:p><w:r><w:t>二、任务完成情况</w:t></w:r></w:p>
    <w:p><w:r><w:t>{{任务完成情况}}</w:t></w:r></w:p>
  </w:body>
</w:document>
""",
    )

    with session_factory() as session:
        file_record = UploadedFile(
            original_name="preview-template.docx",
            stored_name="preview-template.docx",
            file_type="docx",
            file_size=template_path.stat().st_size,
            storage_path=str(template_path),
        )
        session.add(file_record)
        session.flush()
        template = Template(
            name="Preview Template",
            template_type="daily",
            file_id=file_record.id,
            field_config={"fields": ["title", "completed_tasks"]},
        )
        session.add(template)
        session.commit()
        session.refresh(template)
        template_id = template.id

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        response = client.get(f"/api/v1/templates/{template_id}/preview")
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert response.status_code == 200
    preview = response.json()["data"]
    assert preview["source"] == "docx"
    assert preview["preview_mode"] == "source"
    assert "二、任务完成情况" in preview["body"]
    assert preview["raw_placeholders"] == ["title", "任务完成情况"]
    assert preview["fields"] == ["title", "completed_tasks"]
    assert "source-preview-docx" in preview["html"]
    assert "source-placeholder" in preview["html"]


def test_template_preview_endpoint_returns_uploaded_xlsx_style_preview(
    monkeypatch,
    tmp_path,
) -> None:
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    template_path = tmp_path / "uploads" / "preview-template.xlsx"
    template_path.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "日报模板"
    worksheet.merge_cells("A1:C1")
    worksheet["A1"] = "{{title}}"
    worksheet["A1"].font = Font(name="Microsoft YaHei", size=16, bold=True, color="1D4ED8")
    worksheet["A1"].alignment = Alignment(horizontal="center")
    worksheet["A2"] = "二、任务完成情况"
    worksheet["A2"].fill = PatternFill("solid", fgColor="D9EAD3")
    worksheet["A3"] = "{{completed_tasks}}"
    workbook.save(template_path)

    with session_factory() as session:
        file_record = UploadedFile(
            original_name="preview-template.xlsx",
            stored_name="preview-template.xlsx",
            file_type="xlsx",
            file_size=template_path.stat().st_size,
            storage_path=str(template_path),
        )
        session.add(file_record)
        session.flush()
        template = Template(
            name="Excel Preview Template",
            template_type="daily",
            file_id=file_record.id,
            field_config={"fields": ["title", "completed_tasks"]},
        )
        session.add(template)
        session.commit()
        session.refresh(template)
        template_id = template.id

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        response = client.get(f"/api/v1/templates/{template_id}/preview")
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert response.status_code == 200
    preview = response.json()["data"]
    assert preview["source"] == "xlsx"
    assert preview["preview_mode"] == "source"
    assert "二、任务完成情况" in preview["body"]
    assert "source-preview-xlsx" in preview["html"]
    assert "font-weight:700" in preview["html"]
    assert "background-color:#D9EAD3" in preview["html"]


def test_template_upload_accepts_pdf_and_preview_embeds_source(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    template_path = tmp_path / "preview-template.pdf"
    _create_template_pdf(template_path)

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        upload_response = client.post(
            "/api/v1/files/upload",
            files={"file": ("preview-template.pdf", template_path.read_bytes(), "application/pdf")},
        )
        upload = upload_response.json()["data"]
        create_response = client.post(
            "/api/v1/templates",
            json={
                "name": "PDF Preview Template",
                "template_type": "daily",
                "file_path": upload["file_id"],
            },
        )
        template = create_response.json()["data"]
        preview_response = client.get(f"/api/v1/templates/{template['id']}/preview")
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert upload_response.status_code == 200
    assert upload["file_type"] == "pdf"
    assert create_response.status_code == 200
    assert template["field_config"]["fields"] == ["title", "completed_tasks", "owner"]
    assert template["field_config"]["parse"]["raw_content"]["source"] == "pdf"

    assert preview_response.status_code == 200
    preview = preview_response.json()["data"]
    assert preview["source"] == "pdf"
    assert preview["preview_mode"] == "source"
    assert preview["fields"] == ["title", "completed_tasks", "owner"]
    assert "source-preview-pdf" in preview["html"]
    assert "application/pdf;base64" in preview["html"]


def test_template_preview_endpoint_returns_default_body(monkeypatch, tmp_path) -> None:
    monkeypatch.setenv("STORAGE_ROOT", str(tmp_path))
    get_settings.cache_clear()

    engine = create_engine(
        "sqlite+pysqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
        future=True,
    )
    Base.metadata.create_all(engine)
    session_factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)

    with session_factory() as session:
        template = Template(
            name="Default Daily",
            template_type="daily",
            field_config={"fields": ["title", "date", "summary", "completed_tasks"]},
        )
        session.add(template)
        session.commit()
        session.refresh(template)
        template_id = template.id

    def override_get_db() -> Generator[Session, None, None]:
        with session_factory() as session:
            yield session

    app.dependency_overrides[get_db] = override_get_db
    try:
        client = TestClient(app)
        response = client.get(f"/api/v1/templates/{template_id}/preview")
    finally:
        app.dependency_overrides.clear()
        get_settings.cache_clear()

    assert response.status_code == 200
    preview = response.json()["data"]
    assert preview["source"] == "default"
    assert preview["preview_mode"] == "text"
    assert preview["html"] is None
    assert "一、今日总结" in preview["body"]
    assert "{{completed_tasks}}" in preview["body"]


def _create_template_pdf(template_path) -> None:
    pdf = canvas.Canvas(str(template_path), pageCompression=0)
    pdf.drawString(72, 720, "{{title}}")
    pdf.drawString(72, 700, "{{completed_tasks}}")
    pdf.drawString(72, 680, "{{owner}}")
    pdf.save()
